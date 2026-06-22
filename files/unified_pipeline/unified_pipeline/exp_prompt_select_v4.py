"""
exp_prompt_select_v2.py — prompt-selection with SINGLE-PATTERN BASELINES + multi-LLM.

Built on top of exp_prompt_select.py — all existing prompts preserved verbatim.
New additions:
  • 3 single-pattern BASELINES (zeroshot / fewshot / cot) — persona prefix stripped
    from the existing PERSONA_ZS / PERSONA_FS / PERSONA_COT families
  • persona_fewshot and persona_cot WITHOUT meta (were missing as standalone methods)
  • argparse --model / --method for one-at-a-time Colab execution
  • multi-LLM loop
  • RUNS=3 (was 1) to average sampling noise
  • complexity: very_complex -> complex (3 categories)

Ablation reads:
  zeroshot vs persona_zeroshot      -> does persona help?
  persona_X vs persona_X+meta       -> does meta help?
  zeroshot vs fewshot vs cot        -> best base strategy?

Metric: BERTScore-F1. Winner = highest mean per (model, method).

Grounding: persona White et al. 2023; few-shot Brown et al. 2020;
           CoT Wei et al. 2022; meta Suzgun & Kalai 2024.
"""
from __future__ import annotations
import json, random, time, statistics, csv, collections, sys, argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.append(str(ROOT))
import config as C
from generation import llm_clients as LLM
from generation.generate import _clean
from generation.run_meta import meta_refine_prompt
from evaluation import metrics as M
from schema import strip_doc_comments
import openpyxl

# ─────────────────────────── knobs ───────────────────────────────────────
XLSX        = ROOT / "data" / "raw" / "10_fixed_examples_for_prompt_engineering.xlsx"
SHEET       = "\u041b\u0438\u0441\u04421"   # "Лист1"
MODEL_KEYS  = ["deepseek_r1"]               # extend: ["deepseek_r1", "groq", "gemini"]
RUNS        = 3
TEMP        = 0.5
META_ROUNDS = 2
MAX_CODE    = 4000
SEED        = 42

ap = argparse.ArgumentParser()
ap.add_argument("--model",  default=None)
ap.add_argument("--method", default=None)
args = ap.parse_args()

random.seed(SEED)


def _style(l: str) -> str:
    return {"java": "Javadoc (/** ... */)",
            "python": 'a docstring (""" ... """)'}.get(l.lower(), "a doc comment")


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA + ZERO-SHOT  (5 wordings) — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
def _zs0(c, l):
    return ("You are an experienced Java developer writing API documentation. Write a "
            "Javadoc comment for the following Java method. The comment must follow "
            "standard Javadoc conventions: start with a concise summary sentence "
            "describing what the method does, then document each parameter with @param, "
            "the return value with @return (only if the method returns a value), and any "
            "thrown exceptions with @throws. Output ONLY the Javadoc comment block "
            "(/** ... */). Do not repeat the method code, and do not add explanations or "
            "any other text. Method: " + c[:MAX_CODE])

PERSONA_ZS = [
    _zs0,
    lambda c, l: (f"Act as a senior {l} engineer and meticulous API-doc author. Write "
                  f"{_style(l)} for the method, documenting ONLY what the code does. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Take the role of a {l} library maintainer reviewing a pull request. Write "
                  f"the {_style(l)} you would require for this method, documenting only "
                  f"observable behaviour (summary, @param, @return, @throws where they apply). "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Assume the persona of a {l} API technical writer. Generate concise, accurate "
                  f"{_style(l)} for the method below; include a one-sentence summary, @param for "
                  f"each parameter, @return only if it returns a value, and @throws where "
                  f"applicable. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As an expert {l} engineer responsible for public API documentation, write "
                  f"{_style(l)} for the method below, grounded strictly in its code and following "
                  f"standard Javadoc conventions. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# PERSONA + CoT  (5 wordings) — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
PERSONA_COT = [
    lambda c, l: (f"Act as a senior {l} engineer. Reason step by step about the method's single "
                  f"responsibility, its parameters, return value and exceptions; then output ONLY "
                  f"the final {_style(l)} block (hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are an expert {l} API-doc author. Think through what the code does first, "
                  f"then write {_style(l)} from that analysis. Show ONLY the final doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As a careful {l} developer, work out the contract of this method (inputs, "
                  f"output, side effects, thrown exceptions) step by step, then produce ONLY the "
                  f"resulting {_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Act as a {l} reviewer. Reason it through: identify the purpose, each parameter, "
                  f"the return, and any @throws conditions; afterwards output ONLY the final "
                  f"{_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are a meticulous {l} engineer. First derive the method's behaviour by "
                  f"reasoning about its control flow, then write grounded {_style(l)}. Return ONLY "
                  f"the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# FEW-SHOT EXEMPLARS — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
FEWSHOT_EXEMPLARS = [
    {"code": ("public int add(int a, int b) {\n    return a + b;\n}"),
     "doc":  ("/**\n * Returns the sum of two integers.\n *\n * @param a the first addend\n"
              " * @param b the second addend\n * @return the sum of {@code a} and {@code b}\n */")},
    {"code": ("public String get(int index) {\n    if (index < 0 || index >= size)\n"
              "        throw new IndexOutOfBoundsException();\n    return items[index];\n}"),
     "doc":  ("/**\n * Returns the element at the specified position.\n *\n"
              " * @param index the index of the element to return\n"
              " * @return the element at {@code index}\n"
              " * @throws IndexOutOfBoundsException if the index is out of range\n */")},
]

def _fewshot_block(l: str) -> str:
    return "\n\n".join(f"```{l}\n{e['code']}\n```\n{e['doc']}" for e in FEWSHOT_EXEMPLARS)

# ═══════════════════════════════════════════════════════════════════════════
# PERSONA + FEW-SHOT  (5 wordings) — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
PERSONA_FS = [
    lambda c, l: (f"Act as a senior {l} engineer and API-doc author. Here are examples of "
                  f"method -> {_style(l)}:\n\n{_fewshot_block(l)}\n\nNow write {_style(l)} for this "
                  f"method in the same style. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are an expert {l} API technical writer. Study these worked examples, then "
                  f"document the new method consistently with them:\n\n{_fewshot_block(l)}\n\n"
                  f"NEW METHOD:\n```{l}\n{c[:MAX_CODE]}\n```\nOutput ONLY the {_style(l)}."),
    lambda c, l: (f"As a {l} library maintainer, follow the documentation style shown below.\n\n"
                  f"EXAMPLES:\n{_fewshot_block(l)}\n\nDocument this method the same way. Output ONLY "
                  f"the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You write API docs for a {l} project. Given these reference pairs of code and "
                  f"its {_style(l)}:\n\n{_fewshot_block(l)}\n\nProduce {_style(l)} for the following "
                  f"method, matching their level of detail. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Act as a senior {l} engineer. Learn from these examples of well-documented "
                  f"methods:\n\n{_fewshot_block(l)}\n\nApply the same conventions to write "
                  f"{_style(l)} for the method below. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# SINGLE-PATTERN BASELINES — persona prefix stripped from the existing families.
# Each prompt[i] = same instruction body as PERSONA_ZS/FS/COT[i],
# but without the "Act as / You are / As a ..." opening clause.
# This is the ablation baseline: same wording, no role.
# ═══════════════════════════════════════════════════════════════════════════
ZS = [
    # ZS-0: stripped from _zs0 (keep the instruction, remove "You are an experienced...")
    lambda c, l: (
        "Write a Javadoc comment for the following Java method. "
        "Start with a concise summary sentence, then document each parameter with @param, "
        "the return value with @return (only if applicable), and any exceptions with @throws. "
        "Output ONLY the Javadoc comment block (/** ... */). "
        "Do not repeat the method code. Method: " + c[:MAX_CODE]
    ),
    # ZS-1: stripped from PERSONA_ZS[1] ("Act as a senior ... Write ..." -> "Write ...")
    lambda c, l: (f"Write {_style(l)} for the method below, documenting ONLY what the code does. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # ZS-2: stripped from PERSONA_ZS[2] ("Take the role of..." -> "Write...")
    lambda c, l: (f"Write {_style(l)} for this method, documenting only observable behaviour "
                  f"(summary, @param, @return, @throws where they apply). "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # ZS-3: stripped from PERSONA_ZS[3] ("Assume the persona of..." -> "Generate...")
    lambda c, l: (f"Generate concise, accurate {_style(l)} for the method below; include a "
                  f"one-sentence summary, @param for each parameter, @return only if it returns "
                  f"a value, and @throws where applicable. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # ZS-4: stripped from PERSONA_ZS[4] ("As an expert..." -> "Write...")
    lambda c, l: (f"Write {_style(l)} for the method below, grounded strictly in its code and "
                  f"following standard Javadoc conventions. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

FS = [
    # FS-0: stripped from PERSONA_FS[0] ("Act as a senior... Here are examples" -> "Here are examples")
    lambda c, l: (f"Here are examples of method -> {_style(l)}:\n\n{_fewshot_block(l)}\n\n"
                  f"Now write {_style(l)} for this method in the same style. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # FS-1: stripped from PERSONA_FS[1] ("You are an expert... Study..." -> "Study...")
    lambda c, l: (f"Study these worked examples, then document the new method consistently "
                  f"with them:\n\n{_fewshot_block(l)}\n\n"
                  f"NEW METHOD:\n```{l}\n{c[:MAX_CODE]}\n```\nOutput ONLY the {_style(l)}."),
    # FS-2: stripped from PERSONA_FS[2] ("As a... library maintainer..." -> "Follow...")
    lambda c, l: (f"Follow the documentation style shown below.\n\n"
                  f"EXAMPLES:\n{_fewshot_block(l)}\n\nDocument this method the same way. "
                  f"Output ONLY the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # FS-3: stripped from PERSONA_FS[3] ("You write API docs for a... project." -> "Given...")
    lambda c, l: (f"Given these reference pairs of code and {_style(l)}:\n\n{_fewshot_block(l)}\n\n"
                  f"Produce {_style(l)} for the following method, matching their level of detail. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # FS-4: stripped from PERSONA_FS[4] ("Act as a senior... Learn from..." -> "Learn from...")
    lambda c, l: (f"Learn from these examples of well-documented methods:\n\n{_fewshot_block(l)}\n\n"
                  f"Apply the same conventions to write {_style(l)} for the method below. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

COT = [
    # CoT-0: stripped from PERSONA_COT[0] ("Act as a senior... Reason..." -> "Reason...")
    lambda c, l: (f"Reason step by step about the method's single responsibility, its parameters, "
                  f"return value and exceptions; then output ONLY the final {_style(l)} block "
                  f"(hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # CoT-1: stripped from PERSONA_COT[1] ("You are an expert... Think through..." -> "Think through...")
    lambda c, l: (f"Think through what the code does first, then write {_style(l)} from that "
                  f"analysis. Show ONLY the final doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # CoT-2: stripped from PERSONA_COT[2] ("As a careful... work out..." -> "Work out...")
    lambda c, l: (f"Work out the contract of this method (inputs, output, side effects, thrown "
                  f"exceptions) step by step, then produce ONLY the resulting "
                  f"{_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # CoT-3: stripped from PERSONA_COT[3] ("Act as a... reviewer. Reason..." -> "Reason...")
    lambda c, l: (f"Reason through the method: identify the purpose, each parameter, the return, "
                  f"and any @throws conditions; then output ONLY the final "
                  f"{_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    # CoT-4: stripped from PERSONA_COT[4] ("You are a meticulous... First derive..." -> "First derive...")
    lambda c, l: (f"First derive the method's behaviour by reasoning about its control flow, "
                  f"then write grounded {_style(l)}. Return ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# METHOD REGISTRY  (9 methods x 5 prompts = 45 conditions)
# ═══════════════════════════════════════════════════════════════════════════
METHODS: dict[str, tuple[list, bool]] = {
    # single-pattern baselines (no persona) — NEW
    "zeroshot":              (ZS,         False),
    "fewshot":               (FS,         False),
    "cot":                   (COT,        False),
    # + persona — PERSONA_ZS/FS/COT preserved from exp_prompt_select.py
    "persona_zeroshot":      (PERSONA_ZS, False),
    "persona_fewshot":       (PERSONA_FS, False),   # was missing as standalone
    "persona_cot":           (PERSONA_COT,False),   # was missing as standalone
    # + persona + meta — from exp_prompt_select.py
    "persona_zeroshot+meta": (PERSONA_ZS, True),
    "persona_fewshot+meta":  (PERSONA_FS, True),
    "persona_cot+meta":      (PERSONA_COT,True),
}

N_PROMPTS = 5


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════
def load_examples() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"[ERROR] Excel not found at: {XLSX}")
    ws = openpyxl.load_workbook(XLSX)[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["ID"]).value
        if cid is None:
            continue
        comp = str(ws.cell(r, hdr["complexity_category"]).value or "")
        comp = "complex" if comp == "very_complex" else comp
        out.append({
            "pair_id":       str(cid),
            "code_unit":     strip_doc_comments(str(ws.cell(r, hdr["code"]).value or "")),
            "reference_doc": str(ws.cell(r, hdr["comment"]).value or "").strip(),
            "language":      "java",
            "complexity":    comp,
        })
    return out


# ═══════════════════════════════════════════════════════════════════════════
# GENERATION
# ═══════════════════════════════════════════════════════════════════════════
def run_one(model_key: str, examples: list[dict]) -> list[dict]:
    spec = C.MODELS[model_key]
    provider, model = spec["provider"], spec["model"]
    C.TEMPERATURE = TEMP
    call_fn = lambda pr: LLM.call(provider, model, pr)

    sleep_s = (C.SLEEP_GEMINI if provider == "gemini"
               else C.SLEEP_HF  if provider in ("hf", "hf_text")
               else getattr(C, "SLEEP_GROQ", 0.3) if provider == "groq"
               else 0)

    rows: list[dict] = []
    methods_to_run = ({args.method: METHODS[args.method]}
                      if args.method else METHODS)

    for mname, (family, use_meta) in methods_to_run.items():
        for pid in range(N_PROMPTS):
            builder = family[pid]
            for ex in examples:
                base = builder(ex["code_unit"], ex["language"])
                prompt = (meta_refine_prompt(base, call_fn, rounds=META_ROUNDS)
                          if use_meta else base)
                for run in range(RUNS):
                    out = LLM.call(provider, model, prompt)
                    rows.append({
                        "model":         model_key,
                        "method":        mname,
                        "prompt_id":     pid,
                        "pair_id":       ex["pair_id"],
                        "complexity":    ex["complexity"],
                        "run":           run,
                        "temperature":   TEMP,
                        "reference_doc": ex["reference_doc"],
                        "generated":     _clean(out),
                        "status":        "ok" if LLM.is_ok(out) else "fail",
                    })
                    time.sleep(sleep_s)

            # crash-safe incremental save after every (method, prompt_id) block
            suffix = f"_{args.method}" if args.method else ""
            fp = C.DATA_RESULTS / f"prompt_select_v2_{model_key}{suffix}.jsonl"
            fp.parent.mkdir(parents=True, exist_ok=True)
            with open(fp, "w", encoding="utf-8") as f:
                for x in rows:
                    f.write(json.dumps(x, ensure_ascii=False) + "\n")

            done = sum(1 for x in rows
                       if x["model"] == model_key
                       and x["method"] == mname
                       and x["prompt_id"] == pid)
            print(f"[{model_key}] {mname} #{pid} -> {done} gens saved")

    return rows


# ═══════════════════════════════════════════════════════════════════════════
# SCORING + SUMMARY
# ═══════════════════════════════════════════════════════════════════════════
def score(rows: list[dict]) -> None:
    ok = [r for r in rows if r["status"] == "ok" and r["generated"]]
    if not ok:
        print("[WARN] No successful generations to score.")
        return

    f1_scores = M.bertscore_batch(
        [r["reference_doc"] for r in ok],
        [r["generated"]     for r in ok],
    )
    for r, b in zip(ok, f1_scores):
        r["bertscore_f1"] = b

    agg: dict[tuple, list] = collections.defaultdict(list)
    for r in ok:
        if r.get("bertscore_f1") is not None:
            agg[(r["model"], r["method"], r["prompt_id"])].append(r["bertscore_f1"])

    summ = [
        {"model": k[0], "method": k[1], "prompt_id": k[2], "n": len(v),
         "f1_mean": round(statistics.mean(v), 4),
         "f1_sd":   round(statistics.pstdev(v), 4) if len(v) > 1 else 0.0}
        for k, v in agg.items()
    ]
    summ.sort(key=lambda x: (x["model"], x["method"], -x["f1_mean"]))

    out_csv = C.DATA_RESULTS / "prompt_select_v2_summary.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["model", "method", "prompt_id",
                                           "n", "f1_mean", "f1_sd"])
        w.writeheader()
        w.writerows(summ)

    print("\n=== BEST PROMPT PER (model, method) ===")
    best: dict[tuple, dict] = {}
    for s in summ:
        best.setdefault((s["model"], s["method"]), s)
    for (mdl, mth), s in best.items():
        print(f"  {mdl:12s} | {mth:26s} -> #{s['prompt_id']}  "
              f"F1={s['f1_mean']} (±{s['f1_sd']}, n={s['n']})")

    print(f"\nSaved: {out_csv}")
    print("\nAblation reads:")
    print("  zeroshot vs persona_zeroshot      -> effect of persona")
    print("  persona_X vs persona_X+meta       -> effect of meta-refinement")
    print("  zeroshot vs fewshot vs cot        -> best base strategy")


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    examples = load_examples()
    comp_dist = dict(collections.Counter(e["complexity"] for e in examples))
    active_models  = [args.model]  if args.model  else MODEL_KEYS
    active_methods = [args.method] if args.method else list(METHODS)

    total = len(active_models) * len(active_methods) * N_PROMPTS * len(examples) * RUNS
    print(f"{len(examples)} examples {comp_dist}")
    print(f"models={active_models} | methods={active_methods}")
    print(f"N_PROMPTS={N_PROMPTS} | RUNS={RUNS} | TEMP={TEMP} | total gens={total}")

    all_rows: list[dict] = []
    for mk in active_models:
        all_rows += run_one(mk, examples)

    score(all_rows)
