"""
exp_prompt_select_v2.py — prompt-selection with SINGLE-PATTERN BASELINES + multi-LLM.

Place in unified_pipeline/ root.  Run:  python exp_prompt_select_v2.py
Iterate (recommended on Colab): one model and/or one method at a time, e.g.
    python exp_prompt_select_v2.py --model deepseek_r1 --method zeroshot

Implements the meeting feedback (Giammaria):
  • single-pattern BASELINES, no persona:   zeroshot · fewshot · cot
  • + persona:                              persona_zeroshot · persona_fewshot · persona_cot
  • + persona + meta:                       *_+meta   (meta refines the prompt with the SAME LLM)
  So you can read: zeroshot  vs  persona_zeroshot  -> does persona help?
                   X         vs  X+meta            -> does meta help?
  • multiple LLMs (MODEL_KEYS loop)
  • more RUNS (default 3) to average sampling noise
  • complexity merged to 3 categories (very_complex -> complex)

Metric: BERTScore-F1 vs the reference comment. Winner = highest mean per (model, method).
"""
from __future__ import annotations
import json, random, time, statistics, csv, collections, sys, argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.append(str(ROOT))
import config as C
from generation import llm_clients as LLM
from generation.generate import _clean
from generation.run_meta import meta_refine_prompt
from evaluation import metrics as M
from schema import strip_doc_comments
import openpyxl

# ----------------------------- knobs -----------------------------
XLSX        = ROOT / "data" / "raw" / "10_fixed_examples_for_prompt_engineering.xlsx"
SHEET       = "\u041b\u0438\u0441\u04421"          # "Лист1"
MODEL_KEYS  = ["deepseek_r1"]   # add more: ["deepseek_r1", "groq", "qwen_coder"]
RUNS        = 3                 # reruns per (prompt, example), averaged
TEMP        = 0.5
META_ROUNDS = 2
MAX_CODE    = 4000
SEED        = 42

ap = argparse.ArgumentParser()
ap.add_argument("--model", default=None, help="run only this MODEL_KEYS entry")
ap.add_argument("--method", default=None, help="run only this method name")
args = ap.parse_args()

random.seed(SEED)


def _style(l): return {"java": "Javadoc (/** ... */)"}.get(l.lower(), "a doc comment")

# ===== SINGLE-PATTERN BASELINES (no persona) =====
ZS = [
    lambda c, l: f"Write {_style(l)} for the method. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Document the following {l} method with {_style(l)} (summary, @param, @return, @throws where they apply). Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Generate {_style(l)} for this {l} method, grounded strictly in the code. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
]
FS_EX = [
    {"code": "public int add(int a, int b) {\n    return a + b;\n}",
     "doc": "/**\n * Returns the sum of two integers.\n *\n * @param a the first addend\n * @param b the second addend\n * @return the sum of {@code a} and {@code b}\n */"},
    {"code": "public String get(int index) {\n    if (index < 0 || index >= size)\n        throw new IndexOutOfBoundsException();\n    return items[index];\n}",
     "doc": "/**\n * Returns the element at the specified position.\n *\n * @param index the index of the element to return\n * @return the element at {@code index}\n * @throws IndexOutOfBoundsException if the index is out of range\n */"},
]
def _fs_block(l): return "\n\n".join(f"```{l}\n{e['code']}\n```\n{e['doc']}" for e in FS_EX)
FS = [
    lambda c, l: f"Here are examples of method -> {_style(l)}:\n\n{_fs_block(l)}\n\nNow document this method in the same style. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Follow the documentation style of these examples:\n\n{_fs_block(l)}\n\nDocument the following method the same way. Output ONLY the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Given these code -> {_style(l)} pairs:\n\n{_fs_block(l)}\n\nProduce {_style(l)} for the method below, matching their detail. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
]
COT = [
    lambda c, l: f"Reason step by step about the method's responsibility, parameters, return value and exceptions; then output ONLY the final {_style(l)} block (hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Think through what the code does first, then write {_style(l)} from that analysis. Show ONLY the final doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```",
    lambda c, l: f"Work out the method's contract (inputs, output, side effects, exceptions) step by step, then produce ONLY the resulting {_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```",
]

# ===== + PERSONA versions of each =====
PERSONA = "Act as a senior {l} engineer and meticulous API-doc author. "
def _with_persona(fam):
    return [(lambda c, l, f=f: PERSONA.format(l=l) + f(c, l)) for f in fam]
P_ZS, P_FS, P_COT = _with_persona(ZS), _with_persona(FS), _with_persona(COT)

# method name -> (family, apply_meta)
METHODS = {
    "zeroshot":              (ZS,    False),   # baseline, single pattern
    "fewshot":               (FS,    False),
    "cot":                   (COT,   False),
    "persona_zeroshot":      (P_ZS,  False),   # + persona
    "persona_fewshot":       (P_FS,  False),
    "persona_cot":           (P_COT, False),
    "persona_zeroshot+meta": (P_ZS,  True),    # + persona + meta
    "persona_fewshot+meta":  (P_FS,  True),
    "persona_cot+meta":      (P_COT, True),
}


def load_examples():
    ws = openpyxl.load_workbook(XLSX)[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["ID"]).value
        if cid is None:
            continue
        comp = (ws.cell(r, hdr["complexity_category"]).value or "")
        comp = "complex" if comp == "very_complex" else comp   # merge -> 3 categories
        out.append({"pair_id": str(cid),
                    "code_unit": strip_doc_comments(str(ws.cell(r, hdr["code"]).value or "")),
                    "reference_doc": str(ws.cell(r, hdr["comment"]).value or "").strip(),
                    "language": "java", "complexity": comp})
    return out


def run_one(model_key, examples):
    spec = C.MODELS[model_key]
    provider, model = spec["provider"], spec["model"]
    C.TEMPERATURE = TEMP
    call_fn = lambda pr: LLM.call(provider, model, pr)
    rows = []
    methods = {args.method: METHODS[args.method]} if args.method else METHODS
    for mname, (family, use_meta) in methods.items():
        for pid, builder in enumerate(family):
            for ex in examples:
                base = builder(ex["code_unit"], ex["language"])
                prompt = meta_refine_prompt(base, call_fn, rounds=META_ROUNDS) if use_meta else base
                for run in range(RUNS):
                    out = LLM.call(provider, model, prompt)
                    rows.append({"model": model_key, "method": mname, "prompt_id": pid,
                                 "pair_id": ex["pair_id"], "complexity": ex["complexity"],
                                 "run": run, "reference_doc": ex["reference_doc"],
                                 "generated": _clean(out),
                                 "status": "ok" if LLM.is_ok(out) else "fail"})
                    time.sleep(getattr(C, "SLEEP_GROQ", 0) if provider == "groq" else 0)
            print(f"[{model_key}] {mname} #{pid} done")
    fp = C.DATA_RESULTS / f"prompt_select_v2_{model_key}{('_'+args.method) if args.method else ''}.jsonl"
    fp.parent.mkdir(parents=True, exist_ok=True)
    with open(fp, "w", encoding="utf-8") as f:
        for x in rows:
            f.write(json.dumps(x, ensure_ascii=False) + "\n")
    print(f"  wrote {fp}")
    return rows


def score(rows):
    ok = [r for r in rows if r["status"] == "ok" and r["generated"]]
    if not ok:
        print("No successful generations."); return
    f1 = M.bertscore_batch([r["reference_doc"] for r in ok], [r["generated"] for r in ok])
    for r, b in zip(ok, f1):
        r["bertscore_f1"] = b
    agg = collections.defaultdict(list)
    for r in ok:
        if r.get("bertscore_f1") is not None:
            agg[(r["model"], r["method"], r["prompt_id"])].append(r["bertscore_f1"])
    summ = [{"model": k[0], "method": k[1], "prompt_id": k[2], "n": len(v),
             "f1_mean": round(statistics.mean(v), 4),
             "f1_sd": round(statistics.pstdev(v), 4) if len(v) > 1 else 0.0}
            for k, v in agg.items()]
    summ.sort(key=lambda x: (x["model"], x["method"], -x["f1_mean"]))
    out = C.DATA_RESULTS / "prompt_select_v2_summary.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["model", "method", "prompt_id", "n", "f1_mean", "f1_sd"])
        w.writeheader(); w.writerows(summ)
    print("\n=== BEST PROMPT PER (model, method) ===")
    best = {}
    for s in summ:
        best.setdefault((s["model"], s["method"]), s)
    for (mdl, mth), s in best.items():
        print(f"  {mdl:12s} {mth:24s} -> #{s['prompt_id']}  F1={s['f1_mean']} (±{s['f1_sd']}, n={s['n']})")
    print(f"\nWrote {out}\nCompare e.g. zeroshot vs persona_zeroshot to see persona's effect.")


if __name__ == "__main__":
    ex = load_examples()
    print(f"{len(ex)} examples {dict(collections.Counter(e['complexity'] for e in ex))} | "
          f"models={[args.model] if args.model else MODEL_KEYS} | RUNS={RUNS}")
    all_rows = []
    for mk in ([args.model] if args.model else MODEL_KEYS):
        all_rows += run_one(mk, ex)
    score(all_rows)
