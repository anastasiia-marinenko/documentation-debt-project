"""
exp_prompt_select.py — select the BEST PROMPT per prompt-method (screening).

PLACE in unified_pipeline/ root (next to config.py).  Run:  python exp_prompt_select.py

Design — PERSONA pattern fixed; prompt METHOD varied (4 methods kept):
  M1 persona_zeroshot          control: persona, no meta
  M2 persona_zeroshot + meta   self-refinement on the baseline
  M3 persona_fewshot  + meta   in-context examples
  M4 persona_cot      + meta   explicit reasoning
Each method = 5 wordings. Every wording scored on the SAME 10 fixed examples
(from 10_fixed_examples_for_prompt_engineering.xlsx, CodeSearchNet, mixed
complexity), PAIRED, RUNS=5 times at a CONSTANT temperature.
Winner per method = highest mean BERTScore-F1 (report mean +/- sd).

NOTE: all 5 wordings in each method are GENUINE competitors. The zero-shot
prompt #0 is the user's preferred prompt, included verbatim as one fair
candidate; the experiment decides on the data whether it wins.

Grounding: persona/few-shot framing White et al. 2023; few-shot Brown et al.
2020; CoT Wei et al. 2022 / Kojima et al. 2022; meta Suzgun & Kalai 2024.
"""
from __future__ import annotations
import json, random, time, statistics, csv, collections, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.append(str(ROOT))
import config as C
from generation import llm_clients as LLM
from generation.generate import _clean
from generation.run_meta import meta_refine_prompt
from evaluation import metrics as M
from schema import strip_doc_comments

import openpyxl

import os, re

# ----------------------------- knobs -----------------------------
XLSX        = ROOT / "data" / "raw" / "10_fixed_examples_for_prompt_engineering.xlsx"
SHEET       = "\u041b\u0438\u0441\u04421"   # "Лист1"
MODEL_KEY = "deepseek_r1"     # було "groq" deepseek_r1
N_PROMPTS   = 5               # wordings per method
RUNS        = 1               # reruns per (prompt, example), averaged
TEMP        = 0.5             # held CONSTANT across all conditions
META_ROUNDS = 2               # question-refinement rounds (run_meta convention)
MAX_CODE    = 4000            # char cap fed into prompts (ID 355 is ~6.3k chars)
SEED        = 42

ONLY_METHOD = sys.argv[1] if len(sys.argv) > 1 else None
OUT = C.DATA_RESULTS / (f"prompt_selection_{ONLY_METHOD}.jsonl" if ONLY_METHOD else "prompt_selection.jsonl")
SUMMARY = C.DATA_RESULTS / "prompt_selection_summary.csv"



random.seed(SEED)
_spec = C.MODELS[MODEL_KEY]
PROVIDER, MODEL = _spec["provider"], _spec["model"]
C.TEMPERATURE = TEMP                       # threads temp through llm_clients backends
call_fn = lambda pr: LLM.call(PROVIDER, MODEL, pr)


def _style(lang: str) -> str:
    return {"java": "Javadoc (/** ... */)",
            "python": 'a docstring (""" ... """)'}.get(lang.lower(), "a doc comment")


# ============ M1/M2 base: 5 PERSONA ZERO-SHOT wordings ============
# prompt #0 = the user's preferred prompt, verbatim (fair candidate).
def _zs0(c, l):
    return ("You are an experienced Java developer writing API documentation. Write a "
            "Javadoc comment for the following Java method. The comment must follow "
            "standard Javadoc conventions: start with a concise summary sentence "
            "describing what the method does, then document each parameter with @param, "
            "the return value with @return (only if the method returns a value), and any "
            "thrown exceptions with @throws. Output ONLY the Javadoc comment block "
            "(/** ... */). Do not repeat the method code, and do not add explanations or "
            "any other text. Method: " + c[:MAX_CODE])

PERSONA_ZS = [
    _zs0,
    lambda c, l: (f"Act as a senior {l} engineer and meticulous API-doc author. Write "
                  f"{_style(l)} for the method, documenting ONLY what the code does. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Take the role of a {l} library maintainer reviewing a pull request. Write "
                  f"the {_style(l)} you would require for this method, documenting only "
                  f"observable behaviour (summary, @param, @return, @throws where they apply). "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Assume the persona of a {l} API technical writer. Generate concise, accurate "
                  f"{_style(l)} for the method below; include a one-sentence summary, @param for "
                  f"each parameter, @return only if it returns a value, and @throws where "
                  f"applicable. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As an expert {l} engineer responsible for public API documentation, write "
                  f"{_style(l)} for the method below, grounded strictly in its code and following "
                  f"standard Javadoc conventions. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ============ M4 base: 5 PERSONA + CoT wordings ============
PERSONA_COT = [
    lambda c, l: (f"Act as a senior {l} engineer. Reason step by step about the method's single "
                  f"responsibility, its parameters, return value and exceptions; then output ONLY "
                  f"the final {_style(l)} block (hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are an expert {l} API-doc author. Think through what the code does first, "
                  f"then write {_style(l)} from that analysis. Show ONLY the final doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As a careful {l} developer, work out the contract of this method (inputs, "
                  f"output, side effects, thrown exceptions) step by step, then produce ONLY the "
                  f"resulting {_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Act as a {l} reviewer. Reason it through: identify the purpose, each parameter, "
                  f"the return, and any @throws conditions; afterwards output ONLY the final "
                  f"{_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are a meticulous {l} engineer. First derive the method's behaviour by "
                  f"reasoning about its control flow, then write grounded {_style(l)}. Return ONLY "
                  f"the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

# ============ M3 base: 5 PERSONA + FEW-SHOT wordings ============
# Exemplars are canonical and DISJOINT from the 10 evaluation examples (no leakage).
FEWSHOT_EXEMPLARS = [
    {"code": ("public int add(int a, int b) {\n    return a + b;\n}"),
     "doc": ("/**\n * Returns the sum of two integers.\n *\n * @param a the first addend\n"
             " * @param b the second addend\n * @return the sum of {@code a} and {@code b}\n */")},
    {"code": ("public String get(int index) {\n    if (index < 0 || index >= size)\n"
              "        throw new IndexOutOfBoundsException();\n    return items[index];\n}"),
     "doc": ("/**\n * Returns the element at the specified position.\n *\n"
             " * @param index the index of the element to return\n"
             " * @return the element at {@code index}\n"
             " * @throws IndexOutOfBoundsException if the index is out of range\n */")},
]

def _fewshot_block(l):
    return "\n\n".join(f"```{l}\n{e['code']}\n```\n{e['doc']}" for e in FEWSHOT_EXEMPLARS)

PERSONA_FS = [
    lambda c, l: (f"Act as a senior {l} engineer and API-doc author. Here are examples of "
                  f"method -> {_style(l)}:\n\n{_fewshot_block(l)}\n\nNow write {_style(l)} for this "
                  f"method in the same style. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are an expert {l} API technical writer. Study these worked examples, then "
                  f"document the new method consistently with them:\n\n{_fewshot_block(l)}\n\n"
                  f"NEW METHOD:\n```{l}\n{c[:MAX_CODE]}\n```\nOutput ONLY the {_style(l)}."),
    lambda c, l: (f"As a {l} library maintainer, follow the documentation style shown below.\n\n"
                  f"EXAMPLES:\n{_fewshot_block(l)}\n\nDocument this method the same way. Output ONLY "
                  f"the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You write API docs for a {l} project. Given these reference pairs of code and "
                  f"its {_style(l)}:\n\n{_fewshot_block(l)}\n\nProduce {_style(l)} for the following "
                  f"method, matching their level of detail. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Act as a senior {l} engineer. Learn from these examples of well-documented "
                  f"methods:\n\n{_fewshot_block(l)}\n\nApply the same conventions to write "
                  f"{_style(l)} for the method below. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

METHODS = {
    "persona_zeroshot":      {"family": PERSONA_ZS,  "meta": False},
    "persona_zeroshot+meta": {"family": PERSONA_ZS,  "meta": True},
    "persona_fewshot+meta":  {"family": PERSONA_FS,  "meta": True},
    "persona_cot+meta":      {"family": PERSONA_COT, "meta": True},
}


def load_examples():
    """The 10 fixed evaluation examples from the Excel (ID, code, comment, complexity)."""
    if not XLSX.exists():
        raise SystemExit(f"Place the Excel at: {XLSX}")
    ws = openpyxl.load_workbook(XLSX)[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    examples = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["ID"]).value
        if cid is None:
            continue
        code = strip_doc_comments(str(ws.cell(r, hdr["code"]).value or ""))
        doc = str(ws.cell(r, hdr["comment"]).value or "").strip()
        examples.append({
            "pair_id": str(cid),
            "code_unit": code,
            "reference_doc": doc,
            "language": "java",
            "complexity": ws.cell(r, hdr["complexity_category"]).value,
        })
    return examples


def generate(examples):
    rows = []
    items = [(ONLY_METHOD, METHODS[ONLY_METHOD])] if ONLY_METHOD else METHODS.items()
    for mname, m in items:
        for pid in range(N_PROMPTS):
            builder = m["family"][pid]
            for ex in examples:
                code, lang = ex["code_unit"], ex["language"]
                base = builder(code, lang)
                # refine the prompt ONCE per (method, prompt, example); reuse across reruns
                prompt = (meta_refine_prompt(base, call_fn, rounds=META_ROUNDS)
                          if m["meta"] else base)
                for run in range(RUNS):
                    out = LLM.call(PROVIDER, MODEL, prompt)
                    rows.append({
                        "method": mname, "prompt_id": pid, "pair_id": ex["pair_id"],
                        "complexity": ex["complexity"], "run": run, "temperature": TEMP,
                        "reference_doc": ex["reference_doc"],
                        "generated": _clean(out),
                        "status": "ok" if LLM.is_ok(out) else "fail",
                    })
                    time.sleep(C.SLEEP_GROQ)
            OUT.parent.mkdir(parents=True, exist_ok=True)
            with open(OUT, "w", encoding="utf-8") as f:        # crash-safe incremental
                for x in rows:
                    f.write(json.dumps(x, ensure_ascii=False) + "\n")
            done = sum(1 for x in rows if x["method"] == mname and x["prompt_id"] == pid)
            print(f"[{mname}] prompt #{pid} done ({done} gens)")
    return rows


def score_and_select(rows):
    ok = [r for r in rows if r["status"] == "ok" and r["generated"]]
    if not ok:
        raise SystemExit("No successful generations to score.")
    f1 = M.bertscore_batch([r["reference_doc"] for r in ok],
                           [r["generated"] for r in ok])
    for r, b in zip(ok, f1):
        r["bertscore_f1"] = b

    agg = collections.defaultdict(list)
    for r in ok:
        if r.get("bertscore_f1") is not None:
            agg[(r["method"], r["prompt_id"])].append(r["bertscore_f1"])

    summary = [{
        "method": mn, "prompt_id": pid, "n": len(v),
        "bertscore_f1_mean": round(statistics.mean(v), 4),
        "bertscore_f1_sd": round(statistics.pstdev(v), 4) if len(v) > 1 else 0.0,
    } for (mn, pid), v in agg.items()]
    summary.sort(key=lambda x: (x["method"], -x["bertscore_f1_mean"]))

    with open(SUMMARY, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["method", "prompt_id", "n",
                                          "bertscore_f1_mean", "bertscore_f1_sd"])
        w.writeheader()
        w.writerows(summary)

    print("\n=== PER-PROMPT (mean BERTScore-F1) ===")
    for s in summary:
        star = "  <- user prompt" if (s["method"] == "persona_zeroshot" and s["prompt_id"] == 0) else ""
        print(f"  {s['method']:24s} #{s['prompt_id']}  "
              f"{s['bertscore_f1_mean']:.4f} (\u00b1{s['bertscore_f1_sd']:.4f}, n={s['n']}){star}")

    best = {}
    for s in summary:                       # first per method = highest (already sorted)
        best.setdefault(s["method"], s)
    print("\n=== BEST PROMPT PER METHOD ===")
    for mn, s in best.items():
        print(f"  {mn:24s} -> prompt #{s['prompt_id']}  "
              f"F1={s['bertscore_f1_mean']} (\u00b1{s['bertscore_f1_sd']})")
    print(f"\nWrote {SUMMARY}\nNext: confirm the 4 winners on the full 384 set + IRA with Nasser.")


if __name__ == "__main__":
    ex = load_examples()
    comp = collections.Counter(e["complexity"] for e in ex)
    print(f"{len(ex)} fixed examples {dict(comp)} | {len(METHODS)} methods x {N_PROMPTS} "
          f"prompts x {RUNS} runs = {len(METHODS)*N_PROMPTS*len(ex)*RUNS} generations @ temp={TEMP}")
    rows = generate(ex)
    score_and_select(rows)
