"""
exp_prompt_select_v2.py — prompt-selection experiment: 5 prompts per method,
varied persona prefixes, single-pattern baselines + persona + meta layers.

Place in unified_pipeline/ root (next to config.py).

Run (one method at a time, recommended on Colab):
    python exp_prompt_select_v2.py --model deepseek_r1 --method zeroshot
    python exp_prompt_select_v2.py --model deepseek_r1 --method persona_zeroshot
    ... etc.

Design:
  Single-pattern BASELINES (no persona):
    zeroshot · fewshot · cot                          (5 prompts each)
  + PERSONA prefix (5 distinct persona descriptions):
    persona_zeroshot · persona_fewshot · persona_cot  (5 prompts each)
  + PERSONA + META refinement:
    persona_zeroshot+meta · persona_fewshot+meta · persona_cot+meta

  9 methods × 5 prompts = 45 conditions tested on 10 fixed examples × RUNS reruns.

Ablation reads:
  zeroshot vs persona_zeroshot  -> does persona help?
  X        vs X+meta            -> does meta-refinement help?
  zeroshot vs fewshot vs cot    -> which base strategy is strongest?

Metric: BERTScore-F1 vs reference Javadoc. Winner = highest mean per (model, method).

Grounding: persona/few-shot White et al. 2023; few-shot Brown et al. 2020;
           CoT Wei et al. 2022; meta Suzgun & Kalai 2024.
"""
from __future__ import annotations
import json, random, time, statistics, csv, collections, sys, argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.append(str(ROOT))
import config as C
from generation import llm_clients as LLM
from generation.generate import _clean
from generation.run_meta import meta_refine_prompt
from evaluation import metrics as M
from schema import strip_doc_comments
import openpyxl

# ─────────────────────────── knobs ───────────────────────────────────────
XLSX        = ROOT / "data" / "raw" / "10_fixed_examples_for_prompt_engineering.xlsx"
SHEET       = "\u041b\u0438\u0441\u04421"   # "Лист1"
MODEL_KEYS  = ["deepseek_r1"]               # add more keys from config.MODELS
RUNS        = 3                             # reruns per (prompt, example), averaged
TEMP        = 0.5                           # held constant across all conditions
META_ROUNDS = 2                             # meta-refinement rounds
MAX_CODE    = 4000                          # char cap fed into prompts
SEED        = 42

ap = argparse.ArgumentParser()
ap.add_argument("--model",  default=None, help="run only this MODEL_KEYS entry")
ap.add_argument("--method", default=None, help="run only this method name")
args = ap.parse_args()

random.seed(SEED)


def _style(l: str) -> str:
    return {"java": "Javadoc (/** ... */)","python": 'a docstring (""" ... """)'\
            }.get(l.lower(), "a doc comment")


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA PREFIXES — 5 distinct roles, each used across all persona methods
# so the ablation is fair: same role ordering for ZS / FS / CoT
# ═══════════════════════════════════════════════════════════════════════════
PERSONAS = [
    # P0 – senior engineer
    "You are a senior {l} engineer with 10+ years of experience writing API documentation. ",
    # P1 – library maintainer / code-reviewer
    "You are a {l} library maintainer reviewing a pull request. ",
    # P2 – technical writer (documentation specialist)
    "You are an expert {l} API technical writer whose job is producing clear, accurate documentation. ",
    # P3 – open-source contributor familiar with Javadoc conventions
    "You are an experienced open-source {l} contributor who follows strict Javadoc conventions. ",
    # P4 – software architect focused on correctness
    "You are a {l} software architect responsible for the correctness and completeness of public API docs. ",
]

# ═══════════════════════════════════════════════════════════════════════════
# ZERO-SHOT  (5 wordings, without persona prefix)
# ═══════════════════════════════════════════════════════════════════════════
ZS = [
    # ZS-0: minimal direct instruction
    lambda c, l: (
        f"Write {_style(l)} for the method below. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # ZS-1: explicit tag list
    lambda c, l: (
        f"Document the following {l} method with {_style(l)} covering summary, "
        f"@param, @return (if applicable), and @throws (if applicable). "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # ZS-2: grounded-in-code framing
    lambda c, l: (
        f"Generate {_style(l)} for this {l} method, strictly grounded in what the code "
        f"actually does. Do not add information not visible in the code. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # ZS-3: conciseness emphasis
    lambda c, l: (
        f"Write concise, accurate {_style(l)} for the {l} method below. "
        f"Use a one-sentence summary, then @param, @return, @throws where they apply. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # ZS-4: observable-behaviour framing
    lambda c, l: (
        f"Produce {_style(l)} that documents the observable behaviour of the method below "
        f"(what it does, not how). Include all standard Javadoc tags that apply. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
]

# ═══════════════════════════════════════════════════════════════════════════
# FEW-SHOT EXEMPLARS — 2 canonical examples, disjoint from eval set
# ═══════════════════════════════════════════════════════════════════════════
FS_EX = [
    {
        "code": "public int add(int a, int b) {\n    return a + b;\n}",
        "doc":  (
            "/**\n"
            " * Returns the sum of two integers.\n"
            " *\n"
            " * @param a the first addend\n"
            " * @param b the second addend\n"
            " * @return the sum of {@code a} and {@code b}\n"
            " */"
        ),
    },
    {
        "code": (
            "public String get(int index) {\n"
            "    if (index < 0 || index >= size)\n"
            "        throw new IndexOutOfBoundsException();\n"
            "    return items[index];\n"
            "}"
        ),
        "doc": (
            "/**\n"
            " * Returns the element at the specified position.\n"
            " *\n"
            " * @param index the index of the element to return\n"
            " * @return the element at {@code index}\n"
            " * @throws IndexOutOfBoundsException if the index is out of range\n"
            " */"
        ),
    },
]

def _fs_block(l: str) -> str:
    return "\n\n".join(
        f"```{l}\n{e['code']}\n```\n{e['doc']}" for e in FS_EX
    )

# ═══════════════════════════════════════════════════════════════════════════
# FEW-SHOT  (5 wordings, without persona prefix)
# ═══════════════════════════════════════════════════════════════════════════
FS = [
    # FS-0: learn from examples, output only
    lambda c, l: (
        f"Here are examples of {l} method → {_style(l)}:\n\n{_fs_block(l)}\n\n"
        f"Now document this method in the same style. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # FS-1: follow the style explicitly
    lambda c, l: (
        f"Follow the documentation style of these examples:\n\n{_fs_block(l)}\n\n"
        f"Document the following {l} method the same way. "
        f"Output ONLY the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # FS-2: match level of detail
    lambda c, l: (
        f"Given these code → {_style(l)} pairs:\n\n{_fs_block(l)}\n\n"
        f"Produce {_style(l)} for the method below, matching their level of detail. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # FS-3: apply the same conventions
    lambda c, l: (
        f"Study these well-documented {l} methods:\n\n{_fs_block(l)}\n\n"
        f"Apply the same Javadoc conventions to write {_style(l)} for the method below. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # FS-4: reference pairs framing
    lambda c, l: (
        f"Reference pairs of {l} code and {_style(l)}:\n\n{_fs_block(l)}\n\n"
        f"Using these as a guide, write {_style(l)} for the method below. "
        f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
]

# ═══════════════════════════════════════════════════════════════════════════
# CHAIN-OF-THOUGHT  (5 wordings, without persona prefix)
# ═══════════════════════════════════════════════════════════════════════════
COT = [
    # CoT-0: step-by-step, hide reasoning
    lambda c, l: (
        f"Reason step by step about the method's single responsibility, its parameters, "
        f"return value, and exceptions; then output ONLY the final {_style(l)} block "
        f"(hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # CoT-1: analyse then write
    lambda c, l: (
        f"Think through what the code does first, then write {_style(l)} from that analysis. "
        f"Show ONLY the final doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # CoT-2: derive the contract
    lambda c, l: (
        f"Work out the method's contract (inputs, outputs, side effects, exceptions) step by step, "
        f"then produce ONLY the resulting {_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # CoT-3: identify components then document
    lambda c, l: (
        f"First identify: (1) the method's purpose, (2) each parameter's role, "
        f"(3) the return value, (4) any exceptions. "
        f"Then write ONLY the final {_style(l)} based on your analysis.\n\n"
        f"```{l}\n{c[:MAX_CODE]}\n```"
    ),
    # CoT-4: control-flow reasoning
    lambda c, l: (
        f"Derive the method's behaviour by reasoning about its control flow. "
        f"Then write grounded {_style(l)}. Return ONLY the doc block.\n\n"
        f"```{l}\n{c[:MAX_CODE]}\n```"
    ),
]


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA WRAPPERS — prepend persona[i] to base prompt[i]
# Persona index i corresponds to prompt index i so each wording gets a
# distinct persona, making them genuine independent candidates.
# ═══════════════════════════════════════════════════════════════════════════
def _with_persona(base_family: list) -> list:
    """Return a new family where prompt[i] = PERSONAS[i] + base_family[i]."""
    assert len(base_family) == len(PERSONAS), \
        f"Need {len(PERSONAS)} base prompts to match {len(PERSONAS)} personas"
    result = []
    for i, (persona_tmpl, base_fn) in enumerate(zip(PERSONAS, base_family)):
        def _combined(c, l, p=persona_tmpl, b=base_fn):
            return p.format(l=l) + b(c, l)
        result.append(_combined)
    return result

P_ZS  = _with_persona(ZS)
P_FS  = _with_persona(FS)
P_COT = _with_persona(COT)

# ═══════════════════════════════════════════════════════════════════════════
# METHOD REGISTRY
# ═══════════════════════════════════════════════════════════════════════════
METHODS: dict[str, tuple[list, bool]] = {
    # (family of 5 prompt builders, apply_meta)
    "zeroshot":              (ZS,    False),
    "fewshot":               (FS,    False),
    "cot":                   (COT,   False),
    "persona_zeroshot":      (P_ZS,  False),
    "persona_fewshot":       (P_FS,  False),
    "persona_cot":           (P_COT, False),
    "persona_zeroshot+meta": (P_ZS,  True),
    "persona_fewshot+meta":  (P_FS,  True),
    "persona_cot+meta":      (P_COT, True),
}

N_PROMPTS = 5  # must match len(ZS) == len(FS) == len(COT) == len(PERSONAS)
assert all(len(fam) == N_PROMPTS for fam, _ in METHODS.values()), \
    "All method families must have exactly N_PROMPTS entries"


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════
def load_examples() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"[ERROR] Excel not found at: {XLSX}")
    ws = openpyxl.load_workbook(XLSX)[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["ID"]).value
        if cid is None:
            continue
        comp = str(ws.cell(r, hdr["complexity_category"]).value or "")
        comp = "complex" if comp == "very_complex" else comp   # merge to 3 categories
        out.append({
            "pair_id":       str(cid),
            "code_unit":     strip_doc_comments(str(ws.cell(r, hdr["code"]).value or "")),
            "reference_doc": str(ws.cell(r, hdr["comment"]).value or "").strip(),
            "language":      "java",
            "complexity":    comp,
        })
    return out


# ═══════════════════════════════════════════════════════════════════════════
# GENERATION
# ═══════════════════════════════════════════════════════════════════════════
def run_one(model_key: str, examples: list[dict]) -> list[dict]:
    spec = C.MODELS[model_key]
    provider, model = spec["provider"], spec["model"]
    C.TEMPERATURE = TEMP
    call_fn = lambda pr: LLM.call(provider, model, pr)

    sleep_s = (C.SLEEP_GEMINI if provider == "gemini"
               else C.SLEEP_HF if provider in ("hf", "hf_text")
               else getattr(C, "SLEEP_GROQ", 0.3) if provider == "groq"
               else 0)

    rows: list[dict] = []
    methods_to_run = ({args.method: METHODS[args.method]}
                      if args.method else METHODS)

    for mname, (family, use_meta) in methods_to_run.items():
        for pid, builder in enumerate(family):
            # Meta-refine the prompt ONCE per (method, prompt_id, example); reuse across runs
            for ex in examples:
                base_prompt = builder(ex["code_unit"], ex["language"])
                prompt = (meta_refine_prompt(base_prompt, call_fn, rounds=META_ROUNDS)
                          if use_meta else base_prompt)
                for run in range(RUNS):
                    out = LLM.call(provider, model, prompt)
                    rows.append({
                        "model":         model_key,
                        "method":        mname,
                        "prompt_id":     pid,
                        "persona_idx":   pid,           # same index as prompt for transparency
                        "pair_id":       ex["pair_id"],
                        "complexity":    ex["complexity"],
                        "run":           run,
                        "temperature":   TEMP,
                        "reference_doc": ex["reference_doc"],
                        "generated":     _clean(out),
                        "status":        "ok" if LLM.is_ok(out) else "fail",
                    })
                    time.sleep(sleep_s)

            # crash-safe incremental save after every (method, prompt_id) block
            suffix = f"_{args.method}" if args.method else ""
            fp = C.DATA_RESULTS / f"prompt_select_v2_{model_key}{suffix}.jsonl"
            fp.parent.mkdir(parents=True, exist_ok=True)
            with open(fp, "w", encoding="utf-8") as f:
                for x in rows:
                    f.write(json.dumps(x, ensure_ascii=False) + "\n")
            done = sum(1 for x in rows
                       if x["model"] == model_key
                       and x["method"] == mname
                       and x["prompt_id"] == pid)
            print(f"[{model_key}] {mname} #{pid} "
                  f"(persona: '{PERSONAS[pid][:40].format(l='java')}...') "
                  f"-> {done} gens saved")

    return rows


# ═══════════════════════════════════════════════════════════════════════════
# SCORING + SUMMARY
# ═══════════════════════════════════════════════════════════════════════════
def score(rows: list[dict]) -> None:
    ok = [r for r in rows if r["status"] == "ok" and r["generated"]]
    if not ok:
        print("[WARN] No successful generations to score.")
        return

    f1_scores = M.bertscore_batch(
        [r["reference_doc"] for r in ok],
        [r["generated"] for r in ok],
    )
    for r, b in zip(ok, f1_scores):
        r["bertscore_f1"] = b

    # aggregate per (model, method, prompt_id)
    agg: dict[tuple, list] = collections.defaultdict(list)
    for r in ok:
        if r.get("bertscore_f1") is not None:
            agg[(r["model"], r["method"], r["prompt_id"])].append(r["bertscore_f1"])

    summ = [
        {
            "model":      k[0],
            "method":     k[1],
            "prompt_id":  k[2],
            "persona":    PERSONAS[k[2]].format(l="java").strip()[:60],
            "n":          len(v),
            "f1_mean":    round(statistics.mean(v), 4),
            "f1_sd":      round(statistics.pstdev(v), 4) if len(v) > 1 else 0.0,
        }
        for k, v in agg.items()
    ]
    summ.sort(key=lambda x: (x["model"], x["method"], -x["f1_mean"]))

    out_csv = C.DATA_RESULTS / "prompt_select_v2_summary.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["model", "method", "prompt_id", "persona",
                                           "n", "f1_mean", "f1_sd"])
        w.writeheader()
        w.writerows(summ)

    print("\n=== BEST PROMPT PER (model, method) ===")
    best: dict[tuple, dict] = {}
    for s in summ:
        best.setdefault((s["model"], s["method"]), s)

    for (mdl, mth), s in best.items():
        print(f"  {mdl:12s} | {mth:24s} -> prompt #{s['prompt_id']} "
              f"F1={s['f1_mean']} (±{s['f1_sd']}, n={s['n']})")
        print(f"             | persona: {s['persona']}")

    print(f"\nSaved: {out_csv}")
    print("Ablation reads:")
    print("  zeroshot vs persona_zeroshot      -> effect of persona")
    print("  persona_X vs persona_X+meta       -> effect of meta-refinement")
    print("  zeroshot vs fewshot vs cot        -> best base strategy")


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    examples = load_examples()
    comp_dist = dict(collections.Counter(e["complexity"] for e in examples))
    active_models = [args.model] if args.model else MODEL_KEYS
    active_methods = [args.method] if args.method else list(METHODS)

    total_gens = len(active_models) * len(active_methods) * N_PROMPTS * len(examples) * RUNS
    print(f"{len(examples)} examples {comp_dist}")
    print(f"models={active_models} | methods={active_methods}")
    print(f"N_PROMPTS={N_PROMPTS} | RUNS={RUNS} | TEMP={TEMP}")
    print(f"Total generations: {total_gens}")
    print(f"Persona prefixes: {len(PERSONAS)} distinct roles")

    all_rows: list[dict] = []
    for mk in active_models:
        all_rows += run_one(mk, examples)

    score(all_rows)
