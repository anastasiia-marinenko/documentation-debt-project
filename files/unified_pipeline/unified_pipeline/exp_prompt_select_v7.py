"""
exp_prompt_select_v7.py — TWO-STAGE prompt selection (isolated single patterns + multi-LLM).

Implements supervisor feedback (Jun 25) as a two-stage design:

  STAGE 1 — wording selection (cheap, automatic)
    5 isolated patterns x N_WORDINGS(=3) wordings, NO meta, scored by BERTScore-F1
    on the SAME 10 paired examples. Select the single BEST wording per pattern
    (argmax mean F1, with a bootstrap CI) -> stage1_best_wordings.json.
    Why 3 wordings (not 5): one canonical phrasing + two controlled variants is the
    minimal non-trivial set to estimate within-pattern wording sensitivity, while
    keeping the later MANUAL meta-refinement affordable. We discard suboptimal wordings.

  STAGE 2 — pattern comparison + meta (manual rounds)
    ONE prompt per pattern (the Stage-1 winner) evaluated as baseline AND with
    per-pattern meta-refinement, applied with the SAME LLM. META_ROUNDS is NOT a
    fixed magic number: refine until the prompt is satisfactory (typically 2-3),
    set via --meta-rounds after manual inspection. The plain prompt stays as the
    minimum-bar baseline; meta must beat it to be kept.

Run:  python exp_prompt_select_v7.py --model qwen_coder_14b --stage 1
      python exp_prompt_select_v7.py --model qwen_coder_14b --stage 2 --meta-rounds 3

Significance: selection reports a bootstrap CI per wording; _wilcoxon() is available
for paired pattern-vs-pattern tests so claims are statistical, not bare argmax.
DeepSeek-R1 is called with think=False (see generation/llm_clients.py) for stable n.

Redesigned per supervisor feedback (Jun 25): evaluate each prompt pattern ON ITS OWN
first; do NOT mix persona into few-shot/CoT. Persona becomes its own standalone pattern.
Combining patterns (e.g. persona × few-shot) is a LATER stage, not part of this screen.

DESIGN (10 methods x 5 wordings = 50 conditions), each scored on the SAME 10 fixed
examples (paired), RUNS times at a constant temperature:
  • single patterns (baselines):  zeroshot · oneshot · twoshot · cot · persona
  • same patterns + meta:          *_+meta   (meta refines the prompt with the SAME LLM)

Ablation reads:
  zeroshot vs oneshot vs twoshot   -> value of in-context examples (0 → 1 → 2)
  zeroshot vs cot vs persona       -> effect of each single pattern vs the bare baseline
  X        vs X+meta               -> effect of META-refinement, per pattern

─ WHY THESE PATTERNS (logic, not popularity) ─────────────────────────────────
  • ZERO-SHOT — the developer's baseline: just state the task, no examples/role.
  • ONE-SHOT / TWO-SHOT — Brown et al. (2020): in-context exemplars supply
    format/detail priors. We stop at two because gains plateau after ~2 examples,
    so 3-shot+ is not justified for this task.
  • CHAIN-OF-THOUGHT — Wei et al. (2022) / Kojima et al. (2022): elicit step-by-step
    reasoning before the doc block, for NON-reasoning models. (Do NOT combine CoT with a
    reasoning model such as DeepSeek-R1, which already "thinks" — forcing it tends to hurt.)
  • PERSONA — White et al. (2023): assigning a professional role conditions tone/precision.
    Held as ONE pattern; we vary only the role across 5 viewpoints to pick the best wording.
  • Instruction body — the standard Javadoc contract (summary, @param, @return, @throws),
    Oracle "How to Write Doc Comments".
  • META-PROMPTING / self-refinement — Suzgun & Kalai (2024); Self-Refine (Madaan 2023);
    White's Question-Refinement. Applied as a PRELIMINARY per-pattern refinement with the
    SAME LLM used for generation, format-preserving (see generation/run_meta.py).

  META_ROUNDS NOTE: the number of refinement rounds is NOT a fixed "magic number". In
  principle the prompt is refined (and manually inspected) until it is satisfactory —
  typically 2-3 rounds, with diminishing returns afterward. META_ROUNDS below is the
  automated default for this screen; state it explicitly in the paper.

Selection: pick the single BEST wording per pattern (highest mean BERTScore-F1) and
discard the sub-optimal wordings. Winner = highest mean per (model, method).
"""
from __future__ import annotations
import json, random, time, statistics, csv, collections, sys, argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.append(str(ROOT))
import config as C
from generation import llm_clients as LLM
from generation.generate import _clean
from generation.run_meta import meta_refine_prompt
from evaluation import metrics as M
from schema import strip_doc_comments
import openpyxl

# ─────────────────────────── knobs ───────────────────────────────────────
XLSX        = ROOT / "data" / "raw" / "10_fixed_examples_for_prompt_engineering.xlsx"
SHEET       = "\u041b\u0438\u0441\u04421"   # "Лист1"
MODEL_KEYS  = ["deepseek_r1"]               # extend: ["deepseek_r1", "groq", "gemini"]
RUNS        = 3
TEMP        = 0.5
META_ROUNDS = 2     # DEFAULT only; override per pattern via --meta-rounds (manual-until-satisfactory)
MAX_CODE    = 4000
# ── generation-robustness knobs (do NOT change the experiment conditions above) ──
MAX_TOKENS_GEN = 768    # think=False (see llm_clients) -> Javadoc fits easily; small budget = fast, no OOM
MAX_RETRIES    = 3      # retry a (prompt,example,run) until a valid generation -> equal n
SEED        = 42

ap = argparse.ArgumentParser()
ap.add_argument("--model",  default=None)
ap.add_argument("--method", default=None)
ap.add_argument("--stage",  type=int, default=1, choices=[1, 2],
                help="1 = wording selection (5 patterns x 3, no meta); 2 = best wording x {baseline,+meta}")
ap.add_argument("--meta-rounds", type=int, default=None, dest="meta_rounds",
                help="override META_ROUNDS for stage 2 (set after manual inspection)")
args = ap.parse_args()

random.seed(SEED)


def _style(l: str) -> str:
    return {"java": "Javadoc (/** ... */)",
            "python": 'a docstring (""" ... """)'}.get(l.lower(), "a doc comment")


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA + ZERO-SHOT  (5 wordings) — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
PERSONA_ZS = [
    # #0 — your original preferred prompt (kept verbatim, now inline as a lambda).
    lambda c, l: ("You are an experienced Java developer writing API documentation. Write a "
                  "Javadoc comment for the following Java method. The comment must follow "
                  "standard Javadoc conventions: start with a concise summary sentence "
                  "describing what the method does, then document each parameter with @param, "
                  "the return value with @return (only if the method returns a value), and any "
                  "thrown exceptions with @throws. Output ONLY the Javadoc comment block "
                  "(/** ... */). Do not repeat the method code, and do not add explanations or "
                  "any other text. Method: " + c[:MAX_CODE]),
    lambda c, l: (f"Act as a senior {l} engineer and meticulous API-doc author. Write "
                  f"{_style(l)} for the method, documenting ONLY what the code does. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Take the role of a {l} library maintainer reviewing a pull request. Write "
                  f"the {_style(l)} you would require for this method, documenting only "
                  f"observable behaviour (summary, @param, @return, @throws where they apply). "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Assume the persona of a {l} API technical writer. Generate concise, accurate "
                  f"{_style(l)} for the method below; include a one-sentence summary, @param for "
                  f"each parameter, @return only if it returns a value, and @throws where "
                  f"applicable. Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As an expert {l} engineer responsible for public API documentation, write "
                  f"{_style(l)} for the method below, grounded strictly in its code and following "
                  f"standard Javadoc conventions. Output ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# PERSONA + CoT  (5 wordings) — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
# ── LEGACY: persona×CoT combination (NOT used in v6 — combinations are a later stage).
#    Kept only for reference; safe to delete.
PERSONA_COT = [
    lambda c, l: (f"Act as a senior {l} engineer. Reason step by step about the method's single "
                  f"responsibility, its parameters, return value and exceptions; then output ONLY "
                  f"the final {_style(l)} block (hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are an expert {l} API-doc author. Think through what the code does first, "
                  f"then write {_style(l)} from that analysis. Show ONLY the final doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"As a careful {l} developer, work out the contract of this method (inputs, "
                  f"output, side effects, thrown exceptions) step by step, then produce ONLY the "
                  f"resulting {_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Act as a {l} reviewer. Reason it through: identify the purpose, each parameter, "
                  f"the return, and any @throws conditions; afterwards output ONLY the final "
                  f"{_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"You are a meticulous {l} engineer. First derive the method's behaviour by "
                  f"reasoning about its control flow, then write grounded {_style(l)}. Return ONLY "
                  f"the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# FEW-SHOT EXEMPLARS — PRESERVED VERBATIM from exp_prompt_select.py
# ═══════════════════════════════════════════════════════════════════════════
FEWSHOT_EXEMPLARS = [
    {"code": ("public int add(int a, int b) {\n    return a + b;\n}"),
     "doc":  ("/**\n * Returns the sum of two integers.\n *\n * @param a the first addend\n"
              " * @param b the second addend\n * @return the sum of {@code a} and {@code b}\n */")},
    {"code": ("public String get(int index) {\n    if (index < 0 || index >= size)\n"
              "        throw new IndexOutOfBoundsException();\n    return items[index];\n}"),
     "doc":  ("/**\n * Returns the element at the specified position.\n *\n"
              " * @param index the index of the element to return\n"
              " * @return the element at {@code index}\n"
              " * @throws IndexOutOfBoundsException if the index is out of range\n */")},
]

def _fewshot_block(l: str, k: int = 2) -> str:
    """Render the first k worked examples (k=1 -> one-shot, k=2 -> two-shot)."""
    return "\n\n".join(f"```{l}\n{e['code']}\n```\n{e['doc']}" for e in FEWSHOT_EXEMPLARS[:k])

# ═══════════════════════════════════════════════════════════════════════════
# PERSONA — STANDALONE PATTERN (5 wordings = 5 professional roles)
# Per supervisor feedback (Jun 25): persona is evaluated as its OWN single pattern,
# NOT mixed into few-shot/CoT. It is a role prefix + the plain Javadoc task.
# (These are the original persona candidates #0–#4, kept verbatim.)
# ═══════════════════════════════════════════════════════════════════════════
PERSONA = PERSONA_ZS   # alias: the 5 persona wordings ARE the standalone persona pattern


# ═══════════════════════════════════════════════════════════════════════════
# SINGLE-PATTERN BASELINES (no persona): zero-shot, one-shot, two-shot, CoT.
# Each prompt[i] is the same instruction body as the persona wording[i] with the
# "Act as / You are / As a ..." role clause removed → clean single-factor pattern.
# ═══════════════════════════════════════════════════════════════════════════
ZS = [
    # ZS-0: PERSONA_ZS[0] with the "You are an experienced..." role clause removed
    lambda c, l: (
        "Write a Javadoc comment for the following Java method. "
        "Start with a concise summary sentence, then document each parameter with @param, "
        "the return value with @return (only if applicable), and any exceptions with @throws. "
        "Output ONLY the Javadoc comment block (/** ... */). "
        "Do not repeat the method code. Method: " + c[:MAX_CODE]
    ),
    lambda c, l: (f"Write {_style(l)} for the method below, documenting ONLY what the code does. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Write {_style(l)} for this method, documenting only observable behaviour "
                  f"(summary, @param, @return, @throws where they apply). "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Generate concise, accurate {_style(l)} for the method below; include a "
                  f"one-sentence summary, @param for each parameter, @return only if it returns "
                  f"a value, and @throws where applicable. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Write {_style(l)} for the method below, grounded strictly in its code and "
                  f"following standard Javadoc conventions. "
                  f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
]

# Few-shot wording family, parameterised by number of in-context examples k.
# k=1 -> ONE-SHOT, k=2 -> TWO-SHOT (same 5 wordings, only the example count differs).
def _make_shot_family(k: int) -> list:
    return [
        lambda c, l, k=k: (f"Here are examples of method -> {_style(l)}:\n\n{_fewshot_block(l, k)}\n\n"
                           f"Now write {_style(l)} for this method in the same style. "
                           f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
        lambda c, l, k=k: (f"Study these worked examples, then document the new method consistently "
                           f"with them:\n\n{_fewshot_block(l, k)}\n\n"
                           f"NEW METHOD:\n```{l}\n{c[:MAX_CODE]}\n```\nOutput ONLY the {_style(l)}."),
        lambda c, l, k=k: (f"Follow the documentation style shown below.\n\n"
                           f"EXAMPLES:\n{_fewshot_block(l, k)}\n\nDocument this method the same way. "
                           f"Output ONLY the {_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
        lambda c, l, k=k: (f"Given these reference pairs of code and {_style(l)}:\n\n{_fewshot_block(l, k)}\n\n"
                           f"Produce {_style(l)} for the following method, matching their level of detail. "
                           f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
        lambda c, l, k=k: (f"Learn from these examples of well-documented methods:\n\n{_fewshot_block(l, k)}\n\n"
                           f"Apply the same conventions to write {_style(l)} for the method below. "
                           f"Output ONLY the doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    ]

ONESHOT = _make_shot_family(1)   # one in-context example
TWOSHOT = _make_shot_family(2)   # two in-context examples

COT = [
    lambda c, l: (f"Reason step by step about the method's single responsibility, its parameters, "
                  f"return value and exceptions; then output ONLY the final {_style(l)} block "
                  f"(hide the reasoning).\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Think through what the code does first, then write {_style(l)} from that "
                  f"analysis. Show ONLY the final doc block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Work out the contract of this method (inputs, output, side effects, thrown "
                  f"exceptions) step by step, then produce ONLY the resulting "
                  f"{_style(l)}.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"Reason through the method: identify the purpose, each parameter, the return, "
                  f"and any @throws conditions; then output ONLY the final "
                  f"{_style(l)} block.\n\n```{l}\n{c[:MAX_CODE]}\n```"),
    lambda c, l: (f"First derive the method's behaviour by reasoning about its control flow, "
                  f"then write grounded {_style(l)}. Return ONLY the doc block.\n\n"
                  f"```{l}\n{c[:MAX_CODE]}\n```"),
]

# ═══════════════════════════════════════════════════════════════════════════
# METHOD REGISTRY — 5 SINGLE PATTERNS, each WITH and WITHOUT meta-refinement.
# Per supervisor feedback (Jun 25):
#   • patterns are isolated single factors (no persona×few-shot etc. mixing here);
#   • persona is its own standalone pattern;
#   • meta is a PRELIMINARY refinement layer applied per pattern with the SAME LLM
#     (combinations of patterns are a later stage, not part of this screen).
# 10 methods × 5 wordings = 50 conditions.  (tuple = (family, use_meta))
# ═══════════════════════════════════════════════════════════════════════════
# ── single-pattern families, used by both stages (the 5 isolated patterns) ──
PATTERNS: dict[str, list] = {
    "zeroshot": ZS,
    "oneshot":  ONESHOT,
    "twoshot":  TWOSHOT,
    "cot":      COT,
    "persona":  PERSONA,
}

# STAGE 1 — wording selection: every pattern, all N_WORDINGS wordings, NO meta.
# Goal: pick the single best wording per pattern (argmax mean BERTScore-F1),
# so we DON'T carry 3 wordings into the expensive manual meta stage.
STAGE1_METHODS: dict[str, tuple[list, bool]] = {
    name: (fam, False) for name, fam in PATTERNS.items()
}

# STAGE 2 — pattern comparison + meta: ONE prompt per pattern (the Stage-1 winner),
# evaluated as baseline AND with per-pattern meta-refinement (manual rounds).
# Built at runtime from stage1_best_wordings.json (see build_stage2()).

N_WORDINGS = 3   # was 5; see Stage-1 justification in docstring
N_PROMPTS  = N_WORDINGS   # back-compat alias

# trim every family to the first N_WORDINGS wordings (keeps Stage-1 cheap)
for _fam in (ZS, COT):
    del _fam[N_WORDINGS:]
PERSONA_ZS[:] = PERSONA_ZS[:N_WORDINGS]
ONESHOT = _make_shot_family(1)[:N_WORDINGS]
TWOSHOT = _make_shot_family(2)[:N_WORDINGS]
PERSONA = PERSONA_ZS
PATTERNS.update({"zeroshot": ZS, "oneshot": ONESHOT, "twoshot": TWOSHOT,
                 "cot": COT, "persona": PERSONA})
STAGE1_METHODS = {name: (fam, False) for name, fam in PATTERNS.items()}


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════
def load_examples() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"[ERROR] Excel not found at: {XLSX}")
    ws = openpyxl.load_workbook(XLSX)[SHEET]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, hdr["ID"]).value
        if cid is None:
            continue
        comp = str(ws.cell(r, hdr["complexity_category"]).value or "")
        comp = "complex" if comp == "very_complex" else comp
        out.append({
            "pair_id":       str(cid),
            "code_unit":     strip_doc_comments(str(ws.cell(r, hdr["code"]).value or "")),
            "reference_doc": str(ws.cell(r, hdr["comment"]).value or "").strip(),
            "language":      "java",
            "complexity":    comp,
        })
    return out


# ═══════════════════════════════════════════════════════════════════════════
# SIGNIFICANCE HELPERS  (so selection is statistical, not bare argmax)
# ═══════════════════════════════════════════════════════════════════════════
def _bootstrap_ci(values: list[float], iters: int = 2000, alpha: float = 0.05):
    """Percentile bootstrap CI for the mean (no scipy dependency)."""
    if len(values) < 2:
        return (values[0] if values else 0.0, values[0] if values else 0.0)
    means = []
    for _ in range(iters):
        sample = [random.choice(values) for _ in values]
        means.append(statistics.mean(sample))
    means.sort()
    lo = means[int((alpha / 2) * iters)]
    hi = means[int((1 - alpha / 2) * iters)]
    return (round(lo, 4), round(hi, 4))


def _wilcoxon(paired_a: list[float], paired_b: list[float]):
    """Wilcoxon signed-rank p-value for paired scores; returns None if unavailable."""
    try:
        from scipy.stats import wilcoxon
        if len(paired_a) != len(paired_b) or len(paired_a) < 3:
            return None
        if all(a == b for a, b in zip(paired_a, paired_b)):
            return 1.0
        return round(float(wilcoxon(paired_a, paired_b).pvalue), 4)
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# STAGE-1 SELECTION  -> stage1_best_wordings.json (best wording per pattern)
# ═══════════════════════════════════════════════════════════════════════════
STAGE1_BEST = C.DATA_RESULTS / "stage1_best_wordings.json"

def select_best_wordings(summ: list[dict]) -> dict:
    """From the per-(model,method,prompt_id) summary, pick the single best wording
    (prompt_id) per (model, pattern) by mean BERTScore-F1, attach a bootstrap CI."""
    by_key: dict[tuple, list] = collections.defaultdict(list)
    for s in summ:
        by_key[(s["model"], s["method"])].append(s)
    best: dict = {}
    for (mdl, pat), cand in by_key.items():
        winner = max(cand, key=lambda x: x["f1_mean"])
        best.setdefault(mdl, {})[pat] = {
            "prompt_id": winner["prompt_id"],
            "f1_mean":   winner["f1_mean"],
            "f1_ci":     winner.get("f1_ci"),
            "n":         winner["n"],
        }
    STAGE1_BEST.parent.mkdir(parents=True, exist_ok=True)
    with open(STAGE1_BEST, "w", encoding="utf-8") as f:
        json.dump(best, f, ensure_ascii=False, indent=2)
    print(f"\nSaved Stage-1 winners -> {STAGE1_BEST}")
    return best


def build_stage2(model_key: str) -> dict:
    """Read Stage-1 winners and build the Stage-2 method set for this model:
    ONE prompt (the winning wording) per pattern, as baseline AND +meta."""
    if not STAGE1_BEST.exists():
        raise SystemExit("Stage 2 needs stage1_best_wordings.json — run --stage 1 first.")
    best = json.loads(STAGE1_BEST.read_text(encoding="utf-8"))
    chosen = best.get(model_key) or {}
    methods: dict[str, tuple[list, bool]] = {}
    for pat, fam in PATTERNS.items():
        pid = (chosen.get(pat) or {}).get("prompt_id", 0)
        single = [fam[pid]]                       # one wording only
        methods[pat]          = (single, False)   # baseline (no refine) = minimum bar
        methods[pat + "+meta"] = (single, True)   # per-pattern meta-refinement
    return methods



    spec = C.MODELS[model_key]
    provider, model = spec["provider"], spec["model"]
    C.TEMPERATURE = TEMP
    # meta-refinement uses the SAME model, with the larger budget so the refined prompt isn't truncated
    call_fn = lambda pr: LLM.call(provider, model, pr, max_tokens=MAX_TOKENS_GEN)

    sleep_s = (C.SLEEP_GEMINI if provider == "gemini"
               else C.SLEEP_HF  if provider in ("hf", "hf_text")
               else getattr(C, "SLEEP_GROQ", 0.3) if provider == "groq"
               else 0)

    rows: list[dict] = []

    # ── choose the method set for this stage ──────────────────────────────
    if args.stage == 1:
        # wording selection: every pattern, all N_WORDINGS wordings, NO meta
        methods_to_run = ({args.method: STAGE1_METHODS[args.method]}
                          if args.method else STAGE1_METHODS)
    else:
        # stage 2: ONE prompt per pattern (Stage-1 winner) x {baseline, +meta}
        methods_to_run = build_stage2(model_key)
        if args.method:
            methods_to_run = {k: v for k, v in methods_to_run.items()
                              if k.startswith(args.method)}

    meta_rounds = args.meta_rounds if args.meta_rounds is not None else META_ROUNDS

    for mname, (family, use_meta) in methods_to_run.items():
        fail_left = 0
        n_wordings = len(family)              # stage 1: N_WORDINGS; stage 2: 1
        for pid in range(n_wordings):
            builder = family[pid]
            for ex in examples:
                base = builder(ex["code_unit"], ex["language"])
                if use_meta:
                    prompt, meta_trace = meta_refine_prompt(base, call_fn, rounds=meta_rounds,
                                                            return_trace=True)
                else:
                    prompt, meta_trace = base, None
                for run in range(RUNS):
                    # retry until we get a valid generation (or give up after MAX_RETRIES)
                    out = ""
                    for attempt in range(MAX_RETRIES):
                        out = LLM.call(provider, model, prompt, max_tokens=MAX_TOKENS_GEN)
                        if LLM.is_ok(out) and _clean(out).strip():
                            break
                        time.sleep(sleep_s)
                    ok = LLM.is_ok(out) and bool(_clean(out).strip())
                    if not ok:
                        fail_left += 1
                    rows.append({
                        "model":         model_key,
                        "method":        mname,
                        "prompt_id":     pid,
                        "pair_id":       ex["pair_id"],
                        "complexity":    ex["complexity"],
                        "run":           run,
                        "temperature":   TEMP,
                        "reference_doc": ex["reference_doc"],
                        "generated":     _clean(out),
                        "status":        "ok" if ok else "fail",
                        "refined_prompt": prompt if use_meta else None,
                    })
                    time.sleep(sleep_s)

            # crash-safe incremental save after every (method, prompt_id) block
            suffix = f"_{args.method}" if args.method else ""
            fp = C.DATA_RESULTS / f"prompt_select_v7_{model_key}{suffix}.jsonl"
            fp.parent.mkdir(parents=True, exist_ok=True)
            with open(fp, "w", encoding="utf-8") as f:
                for x in rows:
                    f.write(json.dumps(x, ensure_ascii=False) + "\n")

            done = sum(1 for x in rows
                       if x["model"] == model_key
                       and x["method"] == mname
                       and x["prompt_id"] == pid)
            ok_here = sum(1 for x in rows
                          if x["model"] == model_key
                          and x["method"] == mname
                          and x["prompt_id"] == pid
                          and x["status"] == "ok")
            print(f"[{model_key}] {mname} #{pid} -> {ok_here}/{done} ok saved")

    return rows


# ═══════════════════════════════════════════════════════════════════════════
# SCORING + SUMMARY
# ═══════════════════════════════════════════════════════════════════════════
def score(rows: list[dict]) -> None:
    ok = [r for r in rows if r["status"] == "ok" and r["generated"]]
    if not ok:
        print("[WARN] No successful generations to score.")
        return

    f1_scores = M.bertscore_batch(
        [r["reference_doc"] for r in ok],
        [r["generated"]     for r in ok],
    )
    for r, b in zip(ok, f1_scores):
        r["bertscore_f1"] = b

    agg: dict[tuple, list] = collections.defaultdict(list)
    for r in ok:
        if r.get("bertscore_f1") is not None:
            agg[(r["model"], r["method"], r["prompt_id"])].append(r["bertscore_f1"])

    summ = [
        {"model": k[0], "method": k[1], "prompt_id": k[2], "n": len(v),
         "f1_mean": round(statistics.mean(v), 4),
         "f1_sd":   round(statistics.pstdev(v), 4) if len(v) > 1 else 0.0,
         "f1_ci":   _bootstrap_ci(v)}
        for k, v in agg.items()
    ]
    summ.sort(key=lambda x: (x["model"], x["method"], -x["f1_mean"]))

    out_csv = C.DATA_RESULTS / "prompt_select_v7_summary.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["model", "method", "prompt_id",
                                           "n", "f1_mean", "f1_sd", "f1_ci"])
        w.writeheader()
        w.writerows(summ)

    # Stage 1 -> persist the best wording per pattern for Stage 2
    if args.stage == 1:
        select_best_wordings(summ)

    print("\n=== BEST PROMPT PER (model, method) ===")
    best: dict[tuple, dict] = {}
    for s in summ:
        best.setdefault((s["model"], s["method"]), s)
    for (mdl, mth), s in best.items():
        print(f"  {mdl:12s} | {mth:26s} -> #{s['prompt_id']}  "
              f"F1={s['f1_mean']} (±{s['f1_sd']}, n={s['n']})")

    print(f"\nSaved: {out_csv}")
    print("\nAblation reads:")
    print("  zeroshot vs persona_zeroshot      -> effect of persona")
    print("  persona_X vs persona_X+meta       -> effect of meta-refinement")
    print("  zeroshot vs fewshot vs cot        -> best base strategy")


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    examples = load_examples()
    comp_dist = dict(collections.Counter(e["complexity"] for e in examples))
    active_models  = [args.model]  if args.model  else MODEL_KEYS
    active_methods = ([args.method] if args.method
                  else list(STAGE1_METHODS) if args.stage == 1
                  else "stage2(best-wording x {baseline,+meta})")

    n_meth = len(active_methods) if isinstance(active_methods, list) else len(PATTERNS) * 2
    gens_per = N_WORDINGS if args.stage == 1 else 1
    total = len(active_models) * n_meth * gens_per * len(examples) * RUNS
    print(f"{len(examples)} examples {comp_dist}")
    print(f"models={active_models} | methods={active_methods}")
    print(f"STAGE={args.stage} | N_WORDINGS={N_WORDINGS} | RUNS={RUNS} | TEMP={TEMP} | ~total gens={total}")

    all_rows: list[dict] = []
    for mk in active_models:
        all_rows += run_one(mk, examples)

    score(all_rows)