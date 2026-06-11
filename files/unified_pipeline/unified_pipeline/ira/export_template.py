"""
ira/export_template.py — turn the IRA sample into a two-rater xlsx
(Anastasiia + Nasser score independently, 0-2 ordinal rubric + overall 0-10).
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
sys.path.append(str(Path(__file__).resolve().parents[1]))
import config as C

ITEMS = ["correct", "useful", "hallucination", "format_correct", "placement", "readability"]


def export(sample_csv=None, out_xlsx=None):
    sample_csv = Path(sample_csv or (C.DATA_RESULTS / "ira_sample.csv"))
    out_xlsx = Path(out_xlsx or (C.DATA_RESULTS / "ira_eval_template.xlsx"))
    df = pd.read_csv(sample_csv)
    wb = Workbook()
    HEAD = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="2F5496")
    WRAP = Alignment(wrap_text=True, vertical="top")
    cols = ["ira_id", "dataset", "repo", "reference_doc", "generated"] + ITEMS + ["overall_0_10", "notes"]

    def sheet(name):
        ws = wb.create_sheet(name); ws.append(cols)
        for j in range(1, len(cols) + 1):
            c = ws.cell(1, j); c.font = HEAD; c.fill = HFILL
        for _, r in df.iterrows():
            ws.append([r["ira_id"], r["dataset"], r["repo"],
                       str(r["reference_doc"]), str(r["generated"])] + [None]*len(ITEMS) + [None, None])
        dv2 = DataValidation(type="whole", operator="between", formula1=0, formula2=2, allow_blank=True)
        dv10 = DataValidation(type="whole", operator="between", formula1=0, formula2=10, allow_blank=True)
        ws.add_data_validation(dv2); ws.add_data_validation(dv10)
        dv2.add(f"F2:K{ws.max_row}"); dv10.add(f"L2:L{ws.max_row}")
        for col, w in {"A":7,"B":16,"C":18,"D":48,"E":48,"M":30}.items():
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row: c.alignment = WRAP
        return ws

    wb.active.title = "README"
    wb["README"]["A1"] = "Rate independently. Items 0-2 (hallucination reverse). overall 0-10. Then run compute_ira.py"
    sheet("rater_Anastasiia"); sheet("rater_Nasser")
    wb.save(out_xlsx)
    print(f"IRA template -> {out_xlsx}")
